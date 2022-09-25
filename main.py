import openpyxl
from openpyxl.styles import PatternFill
import warnings
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.filedialog import askopenfile
root = tk.Tk()
filetypes = (
        [('Excel Files', ('*.xlsx '))])
f1=f2=None

def openA_Excel_file():
    global filetypes,f1
    f1 = askopenfile(filetypes=filetypes)
    if(not(f1==None)):
        openN_show.delete(0, END)
        openN_show.insert(0, f1.name)
def openO_Excel_file():
    global filetypes , f2
    f2 = askopenfile( filetypes=filetypes)
    if(not(f2==None)):
        openO_show.delete(0, END)
        openO_show.insert(0, f2.name)

    # read the text file and show its content on the Text
def program():
    global f2,f1
    if(not(f1==None)):
        n_find=f1.name
        print(n_find)
        o_find=f2.name
        print(o_find)
        saving_obj=n_find.split('/')[-1]
        print(saving_obj)
        warnings.simplefilter(action='ignore', category=UserWarning)
        n_obj = openpyxl.load_workbook(n_find)  # All findings sheet
        sheet1_obj = n_obj.active   #All
        o_obj = openpyxl.load_workbook(o_find)  # old findings sheet
        sheet2_obj = o_obj.active  #old
        row_count = sheet1_obj.max_row
        column_count = sheet1_obj.max_column
        old_name=o_obj.sheetnames  #list of sheets name
        count=0
        row_list=[]
        for i in range(2,row_count+1):
            if not (sheet1_obj.cell(row=i, column=13).value):
                path_obj=None
            else:
             path_obj = sheet1_obj.cell(row=i, column=13).value.split('/')[2] #path
            count += 1
            rule_obj= sheet1_obj.cell(row=i, column=3).value  #rule: RN
            line_obj=sheet1_obj.cell(row=i, column=16).value #line
            function = sheet1_obj.cell(row=i, column=14).value   #.c function
            list=[f'{path_obj}',f'{function}',f'{rule_obj}',f'{line_obj}']
            row_list.append(list)
            print(list)

        for r in row_list:
            if r[0]=='None':
                continue
            spec_sheet = o_obj[r[0]]
            r2 = 0
            col_spec = spec_sheet['A']
            for c in col_spec:
                if not (c.value is None):
                    r2 += 1

            for k in range(17, r2 + 18):

                if ((spec_sheet.cell(row=k, column=3).value == r[1]) and (
                        spec_sheet.cell(row=k, column=4).value == r[2])):
                    if (spec_sheet.cell(row=k, column=5).value == r[3]):

                        for j in range(1, column_count):
                            sheet1_obj.cell(row=row_list.index(r)+2, column=j).fill = PatternFill(fgColor="ff0000", fill_type="solid")
                    elif( sheet1_obj.cell(row=row_list.index(r)+2, column=1).fill != PatternFill(fgColor="ff0000", fill_type="solid")):
                        for j in range(1, column_count):
                            sheet1_obj.cell(row=row_list.index(r)+2, column=j).fill = PatternFill(fgColor="ffff00", fill_type="solid")
        for l in range(2 , len(row_list)):
            if(sheet1_obj.cell(row=l, column=1).fill == PatternFill(fgColor="ff0000", fill_type="solid")):
                for j in range(1, column_count):
                    sheet1_obj.cell(row=l, column=j).fill = PatternFill(fgColor="ffffff", fill_type="solid")
            elif(sheet1_obj.cell(row=l, column=1).fill != PatternFill(fgColor="ff0000", fill_type="solid")
                 and sheet1_obj.cell(row=l, column=1).fill != PatternFill(fgColor="ffff00", fill_type="solid")):
                for j in range(1, column_count):
                    sheet1_obj.cell(row=l, column=j).fill = PatternFill(fgColor="ff0000", fill_type="solid")
        n_obj.save(n_find)
        labeldone.place(x=350,y=90)




root.title("Misra Scanner")
root.configure(bg='light blue')
root.geometry("540x372")
root.resizable(False, False)
tk.Label(root, text="Misra: ", font=('Times', 20), fg="Purple", bg='sky blue').grid(row=0 , column=0)

# Add image file
bg = PhotoImage(file="Brightskies technologies (1).png")

# Show image using label
label1 = Label(root, image=bg)
label1.place(x=0, y=0)
openN_button = ttk.Button(
    root,
    text='Open All Rules File',
    command=openA_Excel_file
)

openN_button.grid(column=0, row=1, sticky='w', padx=10, pady=10)
openN_show = Entry(root)
openN_show.grid(column=1,row=1)
openO_button = ttk.Button(
        root,
    text='Open Reviewed Rules File',
    command=openO_Excel_file
)
openO_show = Entry(root)
openO_show.grid(column=1,row=3)
openO_button.grid(column=0, row=3, sticky='w', padx=10, pady=10)
btn = tk.Button(root, text='Scan files', font=('Times', 12), bg='#a4accc',command=program)
btn.place(x=350,y=50)
labeldone=Label(root,text="Scanning is done")
labelerror=Label(root,text="Note: The scanner only accepts .xlsx extension and workbook type")
label2=Label(root,text="You can change it through file/export/changefile/workbook")
labelerror.place(x=20,y=290)
label2.place(x=20,y=315)
p1=PhotoImage(file='Brightskies technologies (1)-modified.png')
root.iconphoto(False,p1)
root.mainloop()
